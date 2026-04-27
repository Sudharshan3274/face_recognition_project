import face_recognition
import cv2
import numpy as np
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# --- Setup paths ---
IMAGES_PATH = "images"  # folder with known faces
ATTENDANCE_DIR = "attendance_records"  # folder for attendance files
os.makedirs(ATTENDANCE_DIR, exist_ok=True)

# --- Load known faces dynamically ---
known_face_encodings = []
known_face_names = []

print("🔍 Loading known faces from folder:", IMAGES_PATH)

for file_name in os.listdir(IMAGES_PATH):
    if file_name.lower().endswith(('.jpg', '.jpeg', '.png')):
        path = os.path.join(IMAGES_PATH, file_name)
        image = face_recognition.load_image_file(path)
        encodings = face_recognition.face_encodings(image)

        if len(encodings) > 0:
            known_face_encodings.append(encodings[0])
            name = os.path.splitext(file_name)[0]  # remove extension
            known_face_names.append(name)
            print(f"✅ Loaded encoding for {name}")
        else:
            print(f"⚠️ No face found in {file_name}, skipping.")

print("\n✅ All known faces loaded successfully!\n")

# --- Prepare daily attendance file ---
now = datetime.now()
current_date = now.strftime("%Y-%m-%d")
excel_path = os.path.join(ATTENDANCE_DIR, f"{current_date}.xlsx")

# Create new Excel file if not exists
if not os.path.exists(excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Date", "Time"])
    wb.save(excel_path)
else:
    wb = load_workbook(excel_path)
    ws = wb.active

marked_students = set()  # avoid duplicate entries

# --- Start webcam ---
video_capture = cv2.VideoCapture(0)
print("🎥 Starting video... Press 'q' to quit.\n")

while True:
    ret, frame = video_capture.read()
    if not ret:
        print("❌ Failed to capture frame.")
        break

    # Resize and convert frame for faster recognition
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    # Detect faces and encodings
    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
        name = "Unknown"

        if len(face_distances) > 0:
            best_match_index = np.argmin(face_distances)
            if matches[best_match_index]:
                name = known_face_names[best_match_index]

        # Mark attendance if recognized
        if name != "Unknown" and name not in marked_students:
            current_time = datetime.now().strftime("%H:%M:%S")
            ws.append([name, current_date, current_time])
            wb.save(excel_path)
            marked_students.add(name)
            print(f"🟢 {name} marked present at {current_time}")

        # Scale back up face locations
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        # Draw box and name on the frame
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.putText(frame, name, (left, top - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

    # Show video window
    cv2.imshow("Attendance System", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# --- Cleanup ---
video_capture.release()
cv2.destroyAllWindows()
wb.save(excel_path)

print("\n📘 Attendance saved successfully!")